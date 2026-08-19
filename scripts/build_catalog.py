"""Выгрузка из xlsx → catalog/index.json.

Каталог приходит выгрузкой из VK, и данные в нём грязные: поле «Категория»
у части позиций мусорное (у стула стоит «Блузы и рубашки», у дивана —
«Кровати и основания»), поэтому тип мебели выводим из заголовка, а категорию
используем только как запасной вариант.

Признаки («угловой», «еврокнижка», цвет) лежат вперемешку в заголовке и в
свободном тексте описания — единого формата у выгрузки нет.

Фото не скачиваем: ссылки публичные, Wazzup забирает вложение по URL сам.

    python -m scripts.build_catalog "путь/к/каталогу.xlsx"
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

from app.catalog import vocab

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "catalog" / "index.json"

SECTION_RE = re.compile(r"^\s*([А-ЯЁA-Z][^:]{2,40}?)\s*:\s*(.+)$")


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def parse_sections(raw: str) -> dict[str, str]:
    """Метки описания → значения.

    Разбираем ДО схлопывания переносов: без них «Цвет / декор: белый» и
    следующая секция «Характеристики:» сливаются в одну строку, и в цвет
    попадает половина описания.
    """
    out: dict[str, str] = {}
    for line in (raw or "").splitlines():
        m = SECTION_RE.match(line)
        if m:
            key = m.group(1).strip().lower()
            value = m.group(2).strip()
            if value and value not in {"—", "-"} and key not in out:
                out[key] = value[:200]
    return out


def pick_section(sections: dict[str, str], *names: str) -> str:
    for n in names:
        for key, value in sections.items():
            if n in key:
                return value
    return ""


def parse_price(price_cell: str, description: str) -> int:
    for source in (price_cell or "", description or ""):
        m = vocab.PRICE_RE.search(source.replace("\xa0", " "))
        if m:
            digits = re.sub(r"\D", "", m.group(1))
            if digits:
                return int(digits)
    return 0


def build(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # заголовок

    items: list[dict] = []
    for row in rows:
        if not row or not row[1]:
            continue
        title = _clean(str(row[2] or ""))
        description = str(row[3] or "")
        photos = [str(row[i]).strip() for i in range(7, 12)
                  if row[i] and str(row[i]).startswith("http")]
        if not title or not photos:
            continue

        sections = parse_sections(description)
        haystack = f"{title}\n{description}"
        color_text = pick_section(sections, "цвет", "цветовое", "декор")

        items.append({
            "article": f"AV-{str(row[1]).strip()}",
            "external_id": str(row[1]).strip(),
            "name": title,
            "type": vocab.detect_type(title, str(row[5] or "")),
            "category": _clean(str(row[5] or "")),
            "price": parse_price(str(row[4] or ""), description),
            "price_text": _clean(str(row[4] or "")),
            # Цвет сначала из явного поля: там он точнее, чем угаданный по
            # всему тексту, где мелькают цвета из перечня вариантов
            "colors": vocab.detect_colors(f"{title} {color_text}")
                      or vocab.detect_colors(haystack),
            "color_text": color_text,
            "material": pick_section(sections, "материал", "каркас", "корпус",
                                     "наполнение", "крышка"),
            "features": vocab.detect_features(haystack),
            "mechanism": vocab.detect_mechanism(haystack),
            "sizes": vocab.SIZE_RE.findall(description)[:6],
            "description": _clean(description)[:1200],
            "photos": photos,
        })
    return items


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    xlsx = Path(sys.argv[1])
    if not xlsx.is_file():
        print(f"Не найден файл: {xlsx}")
        return 1

    items = build(xlsx)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"Позиций: {len(items)}  →  {OUT}")
    print(f"Фото всего: {sum(len(i['photos']) for i in items)}")
    print("Типы:")
    for t, n in Counter(i["type"] for i in items).most_common():
        print(f"  {n:5}  {t}")
    print(f"Без цвета: {sum(1 for i in items if not i['colors'])}"
          f"   без цены: {sum(1 for i in items if not i['price'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
