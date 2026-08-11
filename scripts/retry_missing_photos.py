"""Докачка недостающих фото каталога + опциональная доиндексация только новых.

Сравнивает URL из Excel с файлами на диске; качает с ретраями.
"""

from __future__ import annotations

import argparse
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import httpx
from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from import_catalog_xlsx import DEFAULT_OUT, DEFAULT_XLSX, UA, _slug  # noqa: E402



def _download_with_retries(
    client: httpx.Client,
    url: str,
    dest: Path,
    max_side: int,
    attempts: int = 4,
) -> bool:
    if dest.exists() and dest.stat().st_size >= 4_000:
        return True
    last_err = ""
    # иногда VK отдаёт 500 на cs=1200 — пробуем чуть меньший размер
    urls = [url]
    if "cs=1200" in url:
        urls.append(url.replace("cs=1200x0", "cs=720x0").replace("cs=1200x", "cs=720x"))
    if "cs=1280" in url:
        urls.append(url.replace("cs=1280x0", "cs=720x0").replace("cs=1280x", "cs=720x"))

    for attempt in range(1, attempts + 1):
        for u in urls:
            try:
                r = client.get(u, timeout=60.0)
                r.raise_for_status()
                dest.parent.mkdir(parents=True, exist_ok=True)
                tmp = dest.with_suffix(dest.suffix + ".part")
                tmp.write_bytes(r.content)
                with Image.open(tmp) as im:
                    im = im.convert("RGB")
                    w, h = im.size
                    long = max(w, h)
                    if long > max_side:
                        scale = max_side / long
                        im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
                    im.save(dest, format="JPEG", quality=90, optimize=True)
                tmp.unlink(missing_ok=True)
                if dest.stat().st_size >= 4_000:
                    return True
            except Exception as exc:
                last_err = str(exc)
                time.sleep(0.4 * attempt)
    print(f"  fail {dest}: {last_err}", flush=True)
    dest.unlink(missing_ok=True)
    Path(str(dest) + ".part").unlink(missing_ok=True)
    return False


def collect_missing(xlsx: Path, out_dir: Path) -> list[tuple[str, str, Path]]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    missing: list[tuple[str, str, Path]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ext_id, title = row[1], row[2]
        if not ext_id or not title:
            continue
        article = _slug(ext_id)
        art_dir = out_dir / article
        photo_urls = [str(u).strip() for u in row[7:12] if u]
        for i, url in enumerate(photo_urls, start=1):
            dest = art_dir / f"{i:02d}.jpg"
            if not dest.exists() or dest.stat().st_size < 4_000:
                missing.append((article, url, dest))
    return missing


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--max-side", type=int, default=1200)
    p.add_argument("--workers", type=int, default=6)
    args = p.parse_args()

    missing = collect_missing(args.xlsx, args.out)
    print(f"missing photos: {len(missing)}", flush=True)
    if not missing:
        return

    ok = fail = 0
    t0 = time.time()

    def work(item: tuple[str, str, Path]) -> bool:
        _art, url, dest = item
        with httpx.Client(headers={"User-Agent": UA}, follow_redirects=True) as client:
            return _download_with_retries(client, url, dest, args.max_side)

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futs = [pool.submit(work, m) for m in missing]
        for i, fut in enumerate(as_completed(futs), start=1):
            if fut.result():
                ok += 1
            else:
                fail += 1
            if i % 5 == 0 or i == len(missing):
                print(f"[{i}/{len(missing)}] ok={ok} fail={fail} {time.time()-t0:.0f}s", flush=True)

    print(f"done ok={ok} fail={fail}", flush=True)


if __name__ == "__main__":
    main()
