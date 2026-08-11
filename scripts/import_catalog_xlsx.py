"""Импорт Excel «А Ваша Мебель» → catalog/<article>/{meta.json, photos}.

Источник: data/catalog_source.xlsx
Фото качаем с URL (VK CDN), до max_side px по длинной стороне.
"""

from __future__ import annotations

import argparse
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import httpx
from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "catalog_source.xlsx"
DEFAULT_OUT = ROOT / "catalog"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)

TYPE_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"кухн", re.I), "кухня"),
    (re.compile(r"диван", re.I), "диван"),
    (re.compile(r"кровать|основан", re.I), "кровать"),
    (re.compile(r"шкаф", re.I), "шкаф"),
    (re.compile(r"стеллаж|полк", re.I), "стеллаж"),
    (re.compile(r"письменн|компьютерн", re.I), "стол письменный"),
    (re.compile(r"\bстол", re.I), "стол"),
    (re.compile(r"тумб", re.I), "тумба"),
    (re.compile(r"комод", re.I), "комод"),
    (re.compile(r"кресл|кушет", re.I), "кресло"),
    (re.compile(r"стул", re.I), "стул"),
    (re.compile(r"пуф", re.I), "пуф"),
    (re.compile(r"обувниц", re.I), "обувница"),
    (re.compile(r"вешалк", re.I), "вешалка"),
    (re.compile(r"зеркал", re.I), "зеркало"),
    (re.compile(r"матрас", re.I), "матрас"),
]

COLOR_RE = re.compile(r"Цвет\s*/\s*декор:\s*(.+)", re.I)
PRICE_RE = re.compile(r"Цена\s+от\s+([\d\s]+)\s*руб", re.I)


def _slug(external_id: str | int) -> str:
    return f"AV-{external_id}"


def _infer_type(title: str, category: str) -> str:
    blob = f"{title} {category}"
    for rx, label in TYPE_RULES:
        if rx.search(blob):
            return label
    cat = (category or "").strip().lower()
    if cat and cat not in {"блузы и рубашки"}:
        return cat
    return "мебель"


def _parse_colors(description: str, title: str) -> list[str]:
    colors: list[str] = []
    m = COLOR_RE.search(description or "")
    raw = m.group(1).strip() if m else ""
    if not raw:
        # иногда цвет в скобках заголовка: «Стул Шейх (орех/золото)»
        m2 = re.search(r"\(([^)]+)\)\s*$", title or "")
        raw = m2.group(1) if m2 else ""
    for part in re.split(r"[/,;]| и ", raw):
        c = part.strip().lower()
        if c and c not in colors and len(c) < 40:
            colors.append(c)
    return colors


def _download_one(
    client: httpx.Client,
    url: str,
    dest: Path,
    max_side: int,
) -> bool:
    if dest.exists() and dest.stat().st_size >= 4_000:
        return True
    try:
        r = client.get(url, timeout=45.0)
        r.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        tmp = dest.with_suffix(dest.suffix + ".part")
        tmp.write_bytes(r.content)
        with Image.open(tmp) as im:
            im = im.convert("RGB")
            w, h = im.size
            long = max(w, h)
            if long > max_side:
                scale = max_side / long
                im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            im.save(dest, format="JPEG", quality=90, optimize=True)
        tmp.unlink(missing_ok=True)
        return dest.stat().st_size >= 4_000
    except Exception as exc:
        print(f"  fail {dest.name}: {exc}")
        dest.unlink(missing_ok=True)
        Path(str(dest) + ".part").unlink(missing_ok=True)
        return False


def import_catalog(
    xlsx: Path,
    out_dir: Path,
    *,
    max_side: int = 1200,
    workers: int = 12,
    limit: int | None = None,
) -> None:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    out_dir.mkdir(parents=True, exist_ok=True)

    jobs: list[tuple[str, str, Path]] = []  # article, url, dest
    metas: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if limit is not None and len(metas) >= limit:
            break
        ext_id, title = row[1], row[2]
        if not ext_id or not title:
            continue
        article = _slug(ext_id)
        description = str(row[3] or "")
        price = str(row[4] or "").strip()
        category = str(row[5] or "").strip()
        photo_urls = [str(u).strip() for u in row[7:12] if u]

        art_dir = out_dir / article
        art_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "article": article,
            "external_id": ext_id,
            "name": str(title).strip(),
            "description": description,
            "price": price,
            "category": category,
            "type": _infer_type(str(title), category),
            "colors": _parse_colors(description, str(title)),
            "photo_count": len(photo_urls),
            "source": "catalog_source.xlsx",
        }
        (art_dir / "meta.json").write_text(
            json.dumps(meta, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        metas[article] = meta
        for i, url in enumerate(photo_urls, start=1):
            dest = art_dir / f"{i:02d}.jpg"
            jobs.append((article, url, dest))

    print(f"артикулов: {len(metas)}, фото к загрузке: {len(jobs)}")
    ok = 0
    fail = 0
    t0 = time.time()

    def _work(item: tuple[str, str, Path]) -> bool:
        _article, url, dest = item
        with httpx.Client(headers={"User-Agent": UA}, follow_redirects=True) as client:
            return _download_one(client, url, dest, max_side)

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(_work, j): j for j in jobs}
        done = 0
        for fut in as_completed(futs):
            done += 1
            if fut.result():
                ok += 1
            else:
                fail += 1
            if done % 100 == 0 or done == len(jobs):
                elapsed = time.time() - t0
                print(f"[{done}/{len(jobs)}] ok={ok} fail={fail} {elapsed:.0f}s")

    index = {
        "articles": len(metas),
        "photos_ok": ok,
        "photos_fail": fail,
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    (out_dir / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"готово: {index}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--max-side", type=int, default=1200)
    p.add_argument("--workers", type=int, default=12)
    p.add_argument("--limit", type=int, default=None, help="ограничить число артикулов")
    args = p.parse_args()
    if not args.xlsx.exists():
        raise SystemExit(f"нет файла: {args.xlsx}")
    import_catalog(
        args.xlsx,
        args.out,
        max_side=args.max_side,
        workers=args.workers,
        limit=args.limit,
    )


if __name__ == "__main__":
    main()
